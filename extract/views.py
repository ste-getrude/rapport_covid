from django.http import HttpResponse

from .models import ExtractRow
from django.shortcuts import render
from django.http.response import JsonResponse

import datetime as dt

import openpyxl

import os
from pathlib import Path

from .functions import is_yesterday_open, write_to_identification_du_cas

# Create your views here.

def ajax_response(request):
    data_list = []
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
#     if request.method == 'GET':
        q = request.GET.get('term', '')
        search_qs = ExtractRow.objects.filter(nom__istartswith=q)
        if search_qs.count() > 0 and q != '':
            for t in search_qs:
                data = {}
#                 data['id'] = t.id
#                 data['value'] = t.nom + ", " + t.prénom + " gr. " + str(t.groupe)
                data[t.id] = {}
                data[t.id]['nom'] = t.nom
                data[t.id]['prénom'] = t.prénom
                data[t.id]['vacciné'] = t.vacciné
                data[t.id]['dob'] =  t.dob
                data[t.id]['groupe'] = t.groupe
                data[t.id]['avisé'] = t.avisé
                data[t.id]['nom_tuteur'] =  t.nom_tuteur
                data[t.id]['prénom_tuteur'] = t.prénom_tuteur 
                data[t.id]['tel'] = t.tel
                data[t.id]['courriel'] = t.courriel 
                data[t.id]['langue'] = t.langue
                data_list.append(data)
        else:
#             data = {0 : 'aucun nom ne commence par '+q}
#             data_list.append(data)
            pass
    else:
        data = {0:"Fail"}
        data_list.append(data)
    data = JsonResponse(data_list, safe=False)
    mimetype = 'application/json'
    return HttpResponse(data, mimetype)

def search_page(request):
    if request.user.is_authenticated:
        return render(request, 'extractrow_form.html')
    else:
        return HttpResponse('Unauthorized', status=401)

def confirmation_view(request):
    if request.method == "POST":
        context = {}
        t0 = dt.date.today()
        t1 = is_yesterday_open(t0)
        t2 = is_yesterday_open(t1)
        t3 = is_yesterday_open(t2)
        
        context['t0'] = t0.strftime("%Y-%m-%d")
        context['t1'] = t1.strftime("%Y-%m-%d")
        context['t2'] = t2.strftime("%Y-%m-%d")
        context['t3'] = t3.strftime("%Y-%m-%d")
        student_id = request.POST['custId']
        student_info = ExtractRow.objects.get(pk=int(student_id))
        sibblings_list = ExtractRow.objects.filter(
                                                nom_tuteur=student_info.nom_tuteur
                                            ).filter(
                                                prénom_tuteur=student_info.prénom_tuteur
                                            ).exclude(
                                                pk = student_id
                                            )
        sibblings_count = ExtractRow.objects.filter(
                                                nom_tuteur=student_info.nom_tuteur
                                            ).filter(
                                                prénom_tuteur=student_info.prénom_tuteur
                                            ).exclude(
                                                pk = student_id
                                            ).count()
        class_mate_list = ExtractRow.objects.filter(groupe = student_info.groupe).exclude(id = student_id)
        symptoms_list = ["fièvre : 38,1 °C (100,6 °F) et plus (température buccale)",
                         "perte soudaine d’odorat sans congestion nasale, avec ou sans perte du goût",
                         "grande fatigue",
                         "perte d’appétit importante",
                         "douleurs musculaires généralisées (non liées à un effort physique)",
                         "mal de tête",
                        "toux (nouvelle ou aggravée)",
                        "essoufflement, difficulté à respirer",
                        "mal de gorge",
                        "nausées",
                        "vomissements",
                        "diarrhée",
                        "maux de ventre"]
        
        context['student_info'] = student_info
        context['symptoms_list'] = symptoms_list
        context['sibblings_list']= sibblings_list
        context['fratterie_nombre']= sibblings_count
        context['class_mate_list']= class_mate_list
        context['nombre_de_collegue'] = len(class_mate_list)
        
    else:
        pass
    
    return render(request, 'extractrow_confirmation.html', context)
        
def excel_view(request):
    if request.method == "POST":
        context = {}
         
        context['child_first_name'] = request.POST['child_first_name']
        context['child_last_name'] = request.POST['child_last_name']
        context['child_groupe'] = request.POST['child_groupe']
        context["apparition_symptômes"] = request.POST["apparition_symptômes"]
        context['passation_test'] = request.POST['passation_test']
        context['date_1'] = request.POST['date_1']
        context['date_2'] = request.POST['date_2']
        context['date_3'] = request.POST['date_3']
        context['diner'] = request.POST['diner']
        context['transport'] = request.POST['transport']
        context['garderie'] = request.POST['garderie']
        context['parascolaire'] = request.POST['parascolaire']
        context['service'] = request.POST['service']
        context['voiturage'] = request.POST['voiturage']
        context['sport-etude'] = request.POST['sport-etude']
        context['equipe_sportive'] = request.POST['equipe_sportive']
        context['symtomes_list'] = ', '.join(request.POST.getlist('symptom_list'))
        context['risque_modere'] = request.POST.getlist('class_mate_list')
        
        
        
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=' + request.POST['child_last_name'] + '_' + request.POST['child_first_name'] + '_école Sainte-Gertrude-gr.' + request.POST['child_groupe'] + '.xlsx'
        
        BASE_DIR = Path(__file__).resolve().parent.parent
        wb = openpyxl.load_workbook(os.path.join(BASE_DIR, 'extract/data/template_covid.xlsx'))
        ws_id_cas = wb['Identification du cas']
        
        
           
           
           
        champ_information_sur_le_cas_dict = {"Nom :" : request.POST['child_last_name'], 
                                             "Prénom :" : request.POST['child_first_name'], 
                                             "Date de naissance :" : request.POST['child_dob'],
                                             "Sexe :" : 'N/A',
                                             'Numéro de téléphone :' : request.POST['child_phone'],
                                             "Adresse complète :" : 'N/A',
                                             "Date d'apparition des symptômes :" : request.POST["apparition_symptômes"],
                                             "Énumérer les symptômes, si connus : " : ', '.join(request.POST.getlist('symptom_list')),
                                             "Numéro du groupe/nom du programme : " : request.POST['child_groupe'],
                                             "Présence de fratrie(s) dans l'école ? (nom, prénom, numéro(s) de groupe) Fratrie #1 :" : request.POST["sibbling_1"],
                                             "Présence de fratrie(s) dans l'école ? (nom, prénom, numéro(s) de groupe) Fratrie #2 :" : request.POST["sibbling_2"],
                                             "Présence de fratrie(s) dans l'école ? (nom, prénom, numéro(s) de groupe) Fratrie #3 :" : request.POST["sibbling_3"],
                                             "Date de passation du test :" : "à fournir",
                                             "Date jour 1 (aaaa-mm-jj)  :" : request.POST['date_1'],
                                             'Date jour 2 (aaaa-mm-jj)\xa0:' : request.POST['date_2'],
                                             "Date jour 3 (aaaa-mm-jj) : \n(dernier jour fréquenté par l'élève)" : request.POST['date_3']
                                            }

        champs_situation_dans_milieu_dict = {'Numéro du groupe/nom du programme : ' : request.POST['child_groupe'],
                                             'Où le cas mange-t-il le midi : ' : request.POST['diner'],
                                             'Transport scolaire \n(préciser numéro du circuit) :  ' : request.POST['transport'],
                                             'Service de garde (groupe) : ' : request.POST['garderie'],
                                             'Activités parascolaires (préciser) : ' : request.POST['parascolaire'],
                                             'Groupe d\'amis réguliers lors des récréations et des pauses :' : request.POST['groupe_amis'],
                                             'Est-ce que le cas reçoit ou donne des services spécialisés :  ' : request.POST['service'],
                                             'Covoiturage : ' : request.POST['voiturage'],
                                             'Sport-Études/Arts-Études :   ' : request.POST['sport-etude'],
                                             'Équipe sportive :  ' : request.POST['equipe_sportive']
                                            }

        champ_informations_sur_le_milieu_dict = {"Nom de l'établissement :" : "École Sainte-Gertrude",
                                                 "Nom du CSS, de la CS ou de l'association, si applicable :" : "CSSPI",
                                                 "Nom de la personne-ressource à l'école :" : "Geneviève Poitras",
                                                 "Coordonnées de la personne-ressource (numéro de téléphone et courriel) :" : "514-328-3566 + genevieve-poitras@csspi.gouv.qc.ca",
                                                 "Type de ventilation: \nmécanique / fenêtres / dispositif de filtration / échangeur d'air / inconnu" : "Fenêtre",
                                                 "Nombre de contacts à risque modéré" : "",
                                                 "Nombre de contacts à risque faible (approximativement)" : ""
                                                }
        
        
        infos_list = [(champ_information_sur_le_cas_dict, 8), 
                     (champs_situation_dans_milieu_dict, 27),
                    (champ_informations_sur_le_milieu_dict, 39)]

        for infos in infos_list:
            write_to_identification_du_cas(infos, ws_id_cas)
        
           
        ws_contacts = wb['Contacts à risque modéré']
        
        row_number = 6
        for student_data_dict in ExtractRow.objects.filter(id__in=request.POST.getlist('class_mate_list')).values():
            ws_contacts.cell(column=2, row=row_number, value=student_data_dict["nom"] + ", " + student_data_dict["prénom"])
            ws_contacts.cell(column=3, row=row_number, value=student_data_dict["vacciné"])
            ws_contacts.cell(column=4, row=row_number, value=student_data_dict["dob"])
            ws_contacts.cell(column=5, row=row_number, value=student_data_dict["groupe"])
            ws_contacts.cell(column=6, row=row_number, value=student_data_dict["avisé"])
            ws_contacts.cell(column=7, row=row_number, value=student_data_dict["nom_tuteur"] + ", " + student_data_dict["prénom_tuteur"])
            ws_contacts.cell(column=8, row=row_number, value=student_data_dict["tel"])
            ws_contacts.cell(column=9, row=row_number, value=student_data_dict["courriel"])
            ws_contacts.cell(column=10, row=row_number, value=student_data_dict["langue"])
            
            row_number +=1
        
        wb.save(response)
            
        
    else:
        pass
    
    #return render(request, 'excel.html', context)
    return response
                
def misa_a_jour_extract(request):
    if request.user.is_authenticated:
        return render(request, 'upload_form.html')
    else:
        return HttpResponse('Unauthorized', status=401)

def mise_a_jour_script(request):
    if request.user.is_authenticated:
        if "GET" == request.method:
            return render(request, 'upload_form.html', {})
        else:
            excel_file = request.FILES["excel_file"]
    
            # you may put validations here to check extension or file size
    
            wb = openpyxl.load_workbook(excel_file)
    
            # getting a particular sheet by name out of many sheets
            
            worksheet = wb.active
#
            # store the data in a dict
            
            row_number = 1
            excel_dic = {}
            for row in worksheet.iter_rows():
                row_data = list()
                for cell in row:
                    if cell.value is not None:
                        row_data.append(str(cell.value))
                excel_dic[row_number] = row_data   
                row_number += 1
                
            # delete all previous record in the database
            
            ExtractRow.objects.all().delete()
            
            # insert the new data
            
            error_msg_list = []
            for k, row in excel_dic.items():
                if k !=1:
                    try:
                        t = ExtractRow(nom = row[0].split(',')[0].strip(), 
                                   prénom = row[0].split(',')[1].strip(), 
                                   vacciné = row[1], 
                                   dob = row[2], 
                                   groupe = row[3], 
                                   avisé = row[4], 
                                   nom_tuteur = row[5].split(',')[0].strip(),  
                                   prénom_tuteur = row[5].split(',')[1].strip(), 
                                   tel = row[6][:14], 
                                   courriel = row[7], 
                                   langue = row[8])
                        t.save()
                    except:
                        error_msg_list.append('Un probème s\'est produit à la ligne ' + str(k))
                        
#            if there were errors

            if len(error_msg_list) > 0:
                return render(request, 'upload_form.html', {"error_message_list" : error_msg_list})
            else:
                return render(request, 'extractrow_form.html')
    else:
        return HttpResponse('Unauthorized', status=401)

def maj_instruction(request): 
    return render(request, 'instructions.html')   
    
    
    
    
    
    
    
    
    
    
    
    